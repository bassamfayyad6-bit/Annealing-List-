import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Final Annealing - Coils Ready", layout="wide")

TITLE_FILL = "1F3864"
TOLERANCE_MM = 0.02  # matching tolerance between L2 exit thickness and Ann target thickness


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def read_any_excel(uploaded_file, sheet_name=0):
    """Read xls or xlsx into a DataFrame, using the right engine."""
    name = uploaded_file.name.lower()
    data = uploaded_file.read()
    uploaded_file.seek(0)
    if name.endswith(".xls"):
        return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine="xlrd")
    return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine="openpyxl")


def normalize_coil(x):
    if pd.isna(x):
        return None
    return str(x).strip().upper()


def find_col(columns, *candidates):
    """Find a column whose name contains all the given substrings (case-insensitive)."""
    cols_lower = {c: str(c).lower() for c in columns}
    for col, low in cols_lower.items():
        if all(cand.lower() in low for cand in candidates):
            return col
    return None


def load_l2_report(uploaded_file):
    df = read_any_excel(uploaded_file, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_coil = find_col(df.columns, "Coil No.") or find_col(df.columns, "Coil No")
    col_pass = find_col(df.columns, "Pass No")
    col_entry = find_col(df.columns, "Entry Thickness")
    col_exit = find_col(df.columns, "Exit Thickness")
    col_end = find_col(df.columns, "End Time")
    col_speed = find_col(df.columns, "Avg. Mill Speed") or find_col(df.columns, "Avg Mill Speed")
    col_weight = find_col(df.columns, "Measured Weight")
    col_alloy = find_col(df.columns, "Alloy")

    required = [col_coil, col_pass, col_entry, col_exit]
    if any(c is None for c in required):
        raise ValueError(
            "L2 report: couldn't find one of the required columns "
            "(Coil No. / Pass No. / Entry Thickness / Exit Thickness). "
            "Check the file's header row."
        )

    df["_coil_key"] = df[col_coil].apply(normalize_coil)
    df = df.dropna(subset=["_coil_key"])

    # take the LAST pass per coil (highest Pass No.; tie-break by End Time)
    sort_cols = [col_pass] + ([col_end] if col_end else [])
    df_sorted = df.sort_values(sort_cols)
    last_pass = df_sorted.groupby("_coil_key", as_index=False).tail(1).copy()

    last_pass["Entry_mm"] = last_pass[col_entry] / 1000.0
    last_pass["Exit_mm"] = last_pass[col_exit] / 1000.0
    last_pass["_pass_no"] = last_pass[col_pass]
    last_pass["_end_time"] = last_pass[col_end] if col_end else None
    last_pass["_speed"] = last_pass[col_speed] if col_speed else None
    last_pass["_weight_t"] = last_pass[col_weight] if col_weight else None
    last_pass["_alloy"] = last_pass[col_alloy] if col_alloy else None
    last_pass["_coil_no_display"] = last_pass[col_coil]

    return last_pass[
        ["_coil_key", "_coil_no_display", "_pass_no", "Entry_mm", "Exit_mm",
         "_end_time", "_speed", "_weight_t", "_alloy"]
    ]


def load_ann_target_sheet(uploaded_file):
    """Find the sheet inside the Ann workbook that has Targeted Th. + HEAT - S.T columns."""
    name = uploaded_file.name.lower()
    data = uploaded_file.read()
    uploaded_file.seek(0)

    if name.endswith(".xls"):
        xls = pd.ExcelFile(io.BytesIO(data), engine="xlrd")
    else:
        xls = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")

    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        df.columns = [str(c).strip() for c in df.columns]
        col_target = find_col(df.columns, "Targeted Th")
        col_heat = find_col(df.columns, "HEAT", "S.T") or find_col(df.columns, "HEAT - S.T")
        if col_target and col_heat:
            col_coil = find_col(df.columns, "COIL Man") or find_col(df.columns, "Coil")
            col_th = find_col(df.columns, "TH [mm]") or find_col(df.columns, "TH")
            col_customer = find_col(df.columns, "Customer")
            col_alloy = "A" if "A" in df.columns else find_col(df.columns, "Alloy")
            col_temper = find_col(df.columns, "T.T")
            col_width = find_col(df.columns, "Width")
            col_tw = find_col(df.columns, "T.W")
            col_od = find_col(df.columns, "OD")
            col_weight = find_col(df.columns, "Weight")
            col_process = find_col(df.columns, "Process")
            col_next = find_col(df.columns, "NEXT")
            col_pass = find_col(df.columns, "PASS")
            col_trim = find_col(df.columns, "Int", "Trim")
            col_spool = find_col(df.columns, "Steel spool")
            col_pseq = find_col(df.columns, "P.Seq") or find_col(df.columns, "P.Seq.")

            df = df.copy()
            df["_coil_key"] = df[col_coil].apply(normalize_coil)
            df = df.dropna(subset=["_coil_key"])

            out = pd.DataFrame({
                "_coil_key": df["_coil_key"],
                "Customer": df[col_customer] if col_customer else None,
                "Alloy": df[col_alloy] if col_alloy else None,
                "Temper": df[col_temper] if col_temper else None,
                "TH_mm": df[col_th] if col_th else None,
                "Targeted_Th_mm": df[col_target],
                "Width": df[col_width] if col_width else None,
                "TW": df[col_tw] if col_tw else None,
                "OD": df[col_od] if col_od else None,
                "Sched_Weight": df[col_weight] if col_weight else None,
                "Process": df[col_process] if col_process else None,
                "Next": df[col_next] if col_next else None,
                "Pass": df[col_pass] if col_pass else None,
                "Trim": df[col_trim] if col_trim else None,
                "SteelSpool": df[col_spool] if col_spool else None,
                "PSeq": df[col_pseq] if col_pseq else None,
                "Designated_Temp": df[col_heat],
            })
            return out, sheet

    raise ValueError(
        "Couldn't find a sheet in the Ann file with both 'Targeted Th.' and "
        "'HEAT - S.T' columns. Check the file structure."
    )


def find_header_row(ws, must_contain=("COIL Man", "Targeted Th")):
    """Scan the first 10 rows of an openpyxl worksheet for the header row."""
    for r in range(1, 11):
        values = [str(c.value) for c in ws[r] if c.value is not None]
        joined = " | ".join(values)
        if all(any(token.lower() in v.lower() for v in values) for token in must_contain):
            return r
    return None


def load_master_plan(uploaded_file):
    """Load the big Cold Rolling Schedule / Plan master file.
    Returns a per-coil dataframe (last known pass per coil) with Targeted Th.,
    Process, Next, and identifying info. Unlike the small Ann list, this file
    normally has NO temperature column - it's used as the authoritative
    source for target thickness + process routing.
    """
    data = uploaded_file.read()
    uploaded_file.seek(0)
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)

    candidates = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hr = find_header_row(ws)
        if hr:
            candidates.append((sn, hr))

    if not candidates:
        raise ValueError(
            "Couldn't find a sheet in the master plan with 'COIL Man #' and "
            "'Targeted Th.' columns."
        )

    # Prefer a sheet literally named like "CRM" (the active cold-mill plan)
    # over broader/historical sheets (e.g. "Rolling Production Plan").
    crm_candidates = [c for c in candidates if "crm" in c[0].lower()]
    best_sheet, best_header_row = (crm_candidates or candidates)[0]

    wb.close()
    df = pd.read_excel(io.BytesIO(data), sheet_name=best_sheet, header=best_header_row - 1, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    col_coil = find_col(df.columns, "COIL Man")
    col_th = find_col(df.columns, "TH [mm]") or find_col(df.columns, "TH")
    col_target = find_col(df.columns, "Targeted Th")
    col_pass = find_col(df.columns, "PASS")
    col_process = find_col(df.columns, "Process")
    col_next = find_col(df.columns, "NEXT")
    col_customer = find_col(df.columns, "Customer")
    col_alloy = "A" if "A" in df.columns else find_col(df.columns, "Alloy")
    col_temper = find_col(df.columns, "T.T")
    col_width = find_col(df.columns, "Width")
    col_tw = find_col(df.columns, "T.W")
    col_od = find_col(df.columns, "OD")
    col_weight = find_col(df.columns, "Weight")
    col_trim = find_col(df.columns, "Int", "Trim")
    col_spool = find_col(df.columns, "Steel spool")
    col_pseq = find_col(df.columns, "P.Seq") or find_col(df.columns, "P.Seq.")

    df = df.dropna(subset=[col_coil]).copy()
    df["_coil_key"] = df[col_coil].apply(normalize_coil)

    def pass_num(v):
        if pd.isna(v):
            return -1
        m = re.search(r"\d+", str(v))
        return int(m.group()) if m else -1

    df["_pass_num"] = df[col_pass].apply(pass_num) if col_pass else 0
    df_sorted = df.sort_values("_pass_num")
    last_rows = df_sorted.groupby("_coil_key", as_index=False).tail(1)

    out = pd.DataFrame({
        "_coil_key": last_rows["_coil_key"],
        "Customer": last_rows[col_customer] if col_customer else None,
        "Alloy": last_rows[col_alloy] if col_alloy else None,
        "Temper": last_rows[col_temper] if col_temper else None,
        "TH_mm_master": last_rows[col_th] if col_th else None,
        "Targeted_Th_mm_master": last_rows[col_target],
        "Width": last_rows[col_width] if col_width else None,
        "TW": last_rows[col_tw] if col_tw else None,
        "OD": last_rows[col_od] if col_od else None,
        "Sched_Weight": last_rows[col_weight] if col_weight else None,
        "Process": last_rows[col_process] if col_process else None,
        "Next": last_rows[col_next] if col_next else None,
        "Pass": last_rows[col_pass] if col_pass else None,
        "Trim": last_rows[col_trim] if col_trim else None,
        "SteelSpool": last_rows[col_spool] if col_spool else None,
        "PSeq": last_rows[col_pseq] if col_pseq else None,
    })
    return out, best_sheet


INT_ANN_FIXED_TEMP = "360°C - 3 h"  # Int Ann temperature is always fixed, regardless of anything else


FINAL_ANN_TOLERANCE_MM = 0.02  # accepted as a match, but flagged for review if not exact


def classify_coils(l2_df, master_df, ann_df):
    """For every coil that finished a pass at Cold Mill (present in L2),
    classify into Int Ann / Final Ann - With Temp / Final Ann - No Temp.
    A coil counts as reaching its target thickness if the difference is
    within ±0.02mm. Rows where the difference isn't exactly zero are
    flagged (via a '_flag_diff' column) so they can be double-checked.
    Returns (int_ann_df, final_with_temp_df, final_no_temp_df).
    """
    # --- Final Ann: match against the Ann list first (freshest data) ---
    ann_merged = l2_df.merge(ann_df, on="_coil_key", how="inner")
    if not ann_merged.empty:
        def reached_ann_target(row):
            target = row.get("Targeted_Th_mm")
            if pd.isna(target):
                return False
            return round(abs(row["Exit_mm"] - target), 4) <= FINAL_ANN_TOLERANCE_MM

        ann_merged["_reached"] = ann_merged.apply(reached_ann_target, axis=1)
        ann_merged["_flag_diff"] = ann_merged.apply(
            lambda r: pd.notna(r.get("Targeted_Th_mm")) and round(r["Exit_mm"], 2) != round(r["Targeted_Th_mm"], 2),
            axis=1,
        )

        def is_final_ann_next(v):
            if pd.isna(v):
                return False
            v = str(v).lower()
            return "f ann" in v or "final ann" in v

        ann_merged["_is_final"] = ann_merged["Next"].apply(is_final_ann_next)

        def temp_present(v):
            return not (pd.isna(v) or str(v).strip() == "")

        ann_merged["_has_temp"] = ann_merged["Designated_Temp"].apply(temp_present)

        final_with_temp_df = ann_merged[
            ann_merged["_is_final"] & ann_merged["_reached"] & ann_merged["_has_temp"]
        ].copy()
    else:
        final_with_temp_df = ann_merged.copy()

    # --- Master plan: used for Int Ann, and for Final Ann coils that
    # haven't been added to the Ann list yet at all ---
    master_merged = l2_df.merge(master_df, on="_coil_key", how="inner")
    if master_merged.empty:
        empty = master_merged.copy()
        return empty, final_with_temp_df, empty.copy()

    def reached_master_target(row):
        target = row.get("Targeted_Th_mm_master")
        if pd.isna(target):
            return False
        return round(abs(row["Exit_mm"] - target), 4) <= FINAL_ANN_TOLERANCE_MM

    master_merged["_reached"] = master_merged.apply(reached_master_target, axis=1)
    master_merged["_flag_diff"] = master_merged.apply(
        lambda r: pd.notna(r.get("Targeted_Th_mm_master")) and round(r["Exit_mm"], 2) != round(r["Targeted_Th_mm_master"], 2),
        axis=1,
    )

    def next_kind(v):
        if pd.isna(v):
            return None
        v = str(v).lower()
        if "int ann" in v or "intermediate ann" in v:
            return "int_ann"
        if "f ann" in v or "final ann" in v:
            return "final_ann"
        return None

    master_merged["_kind"] = master_merged["Next"].apply(next_kind)

    int_ann_df = master_merged[(master_merged["_kind"] == "int_ann") & (master_merged["_reached"])].copy()
    if not int_ann_df.empty:
        int_ann_df["Designated_Temp"] = INT_ANN_FIXED_TEMP  # always fixed, overrides anything else

    ann_coil_keys = set(ann_df["_coil_key"])
    final_no_temp_df = master_merged[
        (master_merged["_kind"] == "final_ann")
        & (master_merged["_reached"])
        & (~master_merged["_coil_key"].isin(ann_coil_keys))
    ].copy()

    return int_ann_df, final_with_temp_df, final_no_temp_df






def build_ready_table(l2_df, ann_df):
    merged = l2_df.merge(ann_df, on="_coil_key", how="inner")
    if merged.empty:
        return merged

    def is_match(row):
        target = row.get("Targeted_Th_mm")
        if pd.isna(target):
            return False
        # exact match (rounded to 2 decimals to avoid float noise, e.g. 0.5800000001)
        return round(row["Exit_mm"], 2) == round(target, 2)

    merged["_matched"] = merged.apply(is_match, axis=1)
    return merged[merged["_matched"]].copy()


def build_missing_temp_table(l2_df, master_df, ann_df):
    """Coils that (per the master plan) truly reached their final Targeted Th.
    and are headed to Final Annealing, but have NO designated temperature
    recorded in the small Ann list (either missing from it, or blank HEAT-S.T).
    """
    merged = l2_df.merge(master_df, on="_coil_key", how="inner")
    if merged.empty:
        return merged

    def is_match(row):
        target = row.get("Targeted_Th_mm_master")
        if pd.isna(target):
            return False
        return round(row["Exit_mm"], 2) == round(target, 2)

    merged["_matched"] = merged.apply(is_match, axis=1)
    final_coils = merged[merged["_matched"]].copy()
    if final_coils.empty:
        return final_coils

    # Only coils actually routed to Final Annealing next (not INT Trim, CM, etc.)
    def is_final_ann(v):
        if pd.isna(v):
            return False
        v = str(v).lower()
        return "f ann" in v or "final ann" in v

    final_coils = final_coils[final_coils["Next"].apply(is_final_ann)].copy()
    if final_coils.empty:
        return final_coils

    temp_lookup = ann_df[["_coil_key", "Designated_Temp"]].drop_duplicates("_coil_key")
    final_coils = final_coils.merge(temp_lookup, on="_coil_key", how="left")

    def temp_missing(v):
        return pd.isna(v) or str(v).strip() == ""

    final_coils["_temp_missing"] = final_coils["Designated_Temp"].apply(temp_missing)
    return final_coils[final_coils["_temp_missing"]].copy()


ANNEALING_HEADERS = [
    "NO", "Customer", "HEAT - S.T", "COIL Man #", "HM thickness",
    "final thickness", "A", "T.T", "Current thickness", "Width", "T.W",
    "Int + Final Trim", "OD [mm]", "Weight [Kg]", "Targeted Th.",
    "Steel spool", "PASS", "Process", "NEXT", "Processed", "P.Seq.",
    "RR", "B.R.R", "CRR",
]
ANNEALING_COL_WIDTHS = [6, 26, 12, 14, 12, 12, 8, 8, 14, 9, 9, 12, 11, 12, 12, 11, 8, 10, 10, 10, 8, 8, 8, 8]


def _write_annealing_sheet(ws, df):
    """Writes one dataframe into a worksheet using the plant's own
    'ANNEALING' sheet layout exactly (see ANNEALING_HEADERS). Fields we don't
    have a data source for (HM thickness, final thickness, Processed, RR,
    B.R.R, CRR) are left blank rather than guessed.
    """
    hdr_row = 1
    ws.append(ANNEALING_HEADERS)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    flag_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow - "please double-check"

    for c in range(1, len(ANNEALING_HEADERS) + 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    r = hdr_row + 1
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        values = [
            i,
            row.get("Customer"),
            row.get("Designated_Temp"),
            row["_coil_no_display"],
            None,  # HM thickness - no data source
            None,  # final thickness - no data source
            row.get("Alloy"),
            row.get("Temper"),
            round(row["Exit_mm"], 2),
            row.get("Width"),
            row.get("TW"),
            row.get("Trim"),
            row.get("OD"),
            row.get("Sched_Weight"),
            row.get("Targeted_Th_mm") if "Targeted_Th_mm" in row else row.get("Targeted_Th_mm_master"),
            row.get("SteelSpool"),
            row.get("Pass") or row.get("_pass_no"),
            row.get("Process"),
            row.get("Next"),
            None,  # Processed - no data source
            row.get("PSeq"),
            None, None, None,  # RR, B.R.R, CRR - no data source
        ]
        ws.append(values)
        for c in range(1, len(ANNEALING_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center
        ws[f"I{r}"].number_format = "0.00"
        ws[f"O{r}"].number_format = "0.00"
        if row.get("_flag_diff"):
            ws[f"I{r}"].fill = flag_fill
            ws[f"O{r}"].fill = flag_fill
        r += 1

    for i, w in enumerate(ANNEALING_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(ANNEALING_HEADERS))}{hdr_row}"
    ws.freeze_panes = f"A{hdr_row + 1}"

    for r_ in ws.iter_rows():
        for c_ in r_:
            c_.font = Font(name="Arial", size=10, bold=c_.font.bold)


def style_workbook(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "ANNEALING"
    _write_annealing_sheet(ws, df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_classification_workbook(int_ann_df, final_with_temp_df, final_no_temp_df):
    """One workbook, three tabs: INT ANN / FINAL ANN - WITH TEMP / FINAL ANN - NO TEMP."""
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("INT ANN")
    _write_annealing_sheet(ws1, int_ann_df)

    ws2 = wb.create_sheet("FINAL ANN - WITH TEMP")
    _write_annealing_sheet(ws2, final_with_temp_df)

    ws3 = wb.create_sheet("FINAL ANN - NO TEMP")
    _write_annealing_sheet(ws3, final_no_temp_df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf




# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.title("🔥 Annealing Routing Tool")
st.caption("Upload the L2 report, the Ann list, and the master Cold Rolling Plan. The tool finds every coil that finished a Cold Mill pass and sorts it into Int Ann / Final Ann - With Temp / Final Ann - No Temp.")
st.caption("Rows highlighted in yellow in the downloaded file have a small thickness difference — please double-check those.")

col1, col2, col3 = st.columns(3)
with col1:
    l2_file = st.file_uploader("L2 Machine Report", type=["xls", "xlsx"])
with col2:
    ann_file = st.file_uploader("Ann List (with Targeted Th. + HEAT - S.T)", type=["xls", "xlsx"])
with col3:
    master_file = st.file_uploader("Cold Rolling Plan (Master)", type=["xls", "xlsx"])

if l2_file and ann_file and master_file:
    if st.button("Process", type="primary"):
        try:
            with st.spinner("Reading L2 report..."):
                l2_df = load_l2_report(l2_file)
            with st.spinner("Reading Ann list..."):
                ann_df, ann_sheet = load_ann_target_sheet(ann_file)
            with st.spinner("Reading master Cold Rolling Plan..."):
                master_df, master_sheet = load_master_plan(master_file)
            st.success(f"Ann sheet: **{ann_sheet}**  |  Master sheet: **{master_sheet}**")

            int_ann_df, final_with_temp_df, final_no_temp_df = classify_coils(l2_df, master_df, ann_df)

            def show_table(df):
                if df.empty:
                    return df
                target_col = "Targeted_Th_mm" if "Targeted_Th_mm" in df.columns else "Targeted_Th_mm_master"
                cols = {
                    "_coil_no_display": "Coil No.", "Customer": "Customer", "Alloy": "Alloy",
                    "Temper": "Temper", "Exit_mm": "Reached Th. (mm)",
                    target_col: "Final Targeted Th. (mm)",
                }
                if "Designated_Temp" in df.columns:
                    cols["Designated_Temp"] = "Temp"
                out = df[list(cols.keys())].rename(columns=cols)
                if "_flag_diff" in df.columns:
                    out.insert(0, "⚠️", df["_flag_diff"].apply(lambda v: "⚠️" if v else ""))
                return out

            st.divider()
            st.subheader(f"🟠 Int Ann ({len(int_ann_df)})")
            if int_ann_df.empty:
                st.info("No coils currently routed to Intermediate Annealing.")
            else:
                st.dataframe(show_table(int_ann_df), use_container_width=True)

            st.divider()
            st.subheader(f"🟢 Final Ann - With Temp ({len(final_with_temp_df)})")
            if final_with_temp_df.empty:
                st.info("No coils currently ready with a designated temperature.")
            else:
                st.dataframe(show_table(final_with_temp_df), use_container_width=True)

            st.divider()
            st.subheader(f"🔴 Final Ann - No Temp ({len(final_no_temp_df)})")
            if final_no_temp_df.empty:
                st.info("No coils are missing a temperature right now.")
            else:
                st.warning(f"{len(final_no_temp_df)} coil(s) need a temperature assigned")
                st.dataframe(show_table(final_no_temp_df), use_container_width=True)

            st.divider()
            excel_buf = build_classification_workbook(int_ann_df, final_with_temp_df, final_no_temp_df)
            st.download_button(
                "⬇️ Download Excel (3 tabs: Int Ann / Final Ann - With Temp / Final Ann - No Temp)",
                data=excel_buf,
                file_name=f"Annealing_Routing_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_classification",
            )
        except Exception as e:
            st.error(f"Error: {e}")
else:
    st.info("Upload all three files (L2 report, Ann list, and the master Cold Rolling Plan) to start.")
