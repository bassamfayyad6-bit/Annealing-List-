import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Final Annealing - Coils Ready", layout="wide")

TITLE_FILL = "1F3864"
TOLERANCE_MM = 0.02  # matching tolerance between L2 exit thickness and Ann target thickness


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def read_any_excel(uploaded_file, sheet_name=0):
    """Read xls or xlsx into a DataFrame, using the right engine."""
    name = uploaded_file.name.lower()
    data = uploaded_file.read()
    uploaded_file.seek(0)
    if name.endswith(".xls"):
        return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine="xlrd")
    return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine="openpyxl")


def normalize_coil(x):
    if pd.isna(x):
        return None
    return str(x).strip().upper()


def find_col(columns, *candidates):
    """Find a column whose name contains all the given substrings (case-insensitive)."""
    cols_lower = {c: str(c).lower() for c in columns}
    for col, low in cols_lower.items():
        if all(cand.lower() in low for cand in candidates):
            return col
    return None


def load_l2_report(uploaded_file):
    df = read_any_excel(uploaded_file, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_coil = find_col(df.columns, "Coil No.") or find_col(df.columns, "Coil No")
    col_pass = find_col(df.columns, "Pass No")
    col_entry = find_col(df.columns, "Entry Thickness")
    col_exit = find_col(df.columns, "Exit Thickness")
    col_end = find_col(df.columns, "End Time")
    col_speed = find_col(df.columns, "Avg. Mill Speed") or find_col(df.columns, "Avg Mill Speed")
    col_weight = find_col(df.columns, "Measured Weight")
    col_alloy = find_col(df.columns, "Alloy")

    required = [col_coil, col_pass, col_entry, col_exit]
    if any(c is None for c in required):
        raise ValueError(
            "L2 report: couldn't find one of the required columns "
            "(Coil No. / Pass No. / Entry Thickness / Exit Thickness). "
            "Check the file's header row."
        )

    df["_coil_key"] = df[col_coil].apply(normalize_coil)
    df = df.dropna(subset=["_coil_key"])

    # take the LAST pass per coil (highest Pass No.; tie-break by End Time)
    sort_cols = [col_pass] + ([col_end] if col_end else [])
    df_sorted = df.sort_values(sort_cols)
    last_pass = df_sorted.groupby("_coil_key", as_index=False).tail(1).copy()

    last_pass["Entry_mm"] = last_pass[col_entry] / 1000.0
    last_pass["Exit_mm"] = last_pass[col_exit] / 1000.0
    last_pass["_pass_no"] = last_pass[col_pass]
    last_pass["_end_time"] = last_pass[col_end] if col_end else None
    last_pass["_speed"] = last_pass[col_speed] if col_speed else None
    last_pass["_weight_t"] = last_pass[col_weight] if col_weight else None
    last_pass["_alloy"] = last_pass[col_alloy] if col_alloy else None
    last_pass["_coil_no_display"] = last_pass[col_coil]

    return last_pass[
        ["_coil_key", "_coil_no_display", "_pass_no", "Entry_mm", "Exit_mm",
         "_end_time", "_speed", "_weight_t", "_alloy"]
    ]


def load_ann_target_sheet(uploaded_file):
    """Find the sheet inside the Ann workbook that has Targeted Th. + HEAT - S.T columns."""
    name = uploaded_file.name.lower()
    data = uploaded_file.read()
    uploaded_file.seek(0)

    if name.endswith(".xls"):
        xls = pd.ExcelFile(io.BytesIO(data), engine="xlrd")
    else:
        xls = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")

    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        df.columns = [str(c).strip() for c in df.columns]
        col_target = find_col(df.columns, "Targeted Th")
        col_heat = find_col(df.columns, "HEAT", "S.T") or find_col(df.columns, "HEAT - S.T")
        if col_target and col_heat:
            col_coil = find_col(df.columns, "COIL Man") or find_col(df.columns, "Coil")
            col_th = find_col(df.columns, "TH [mm]") or find_col(df.columns, "TH")
            col_customer = find_col(df.columns, "Customer")
            col_alloy = find_col(df.columns, "A") if "A" in df.columns else find_col(df.columns, "Alloy")
            col_temper = find_col(df.columns, "T.T")
            col_width = find_col(df.columns, "Width")
            col_tw = find_col(df.columns, "T.W")
            col_od = find_col(df.columns, "OD")
            col_weight = find_col(df.columns, "Weight")
            col_process = find_col(df.columns, "Process")
            col_next = find_col(df.columns, "NEXT")
            col_pass = find_col(df.columns, "PASS")

            df = df.copy()
            df["_coil_key"] = df[col_coil].apply(normalize_coil)
            df = df.dropna(subset=["_coil_key"])

            out = pd.DataFrame({
                "_coil_key": df["_coil_key"],
                "Customer": df[col_customer] if col_customer else None,
                "Alloy": df[col_alloy] if col_alloy else None,
                "Temper": df[col_temper] if col_temper else None,
                "TH_mm": df[col_th] if col_th else None,
                "Targeted_Th_mm": df[col_target],
                "Width": df[col_width] if col_width else None,
                "TW": df[col_tw] if col_tw else None,
                "OD": df[col_od] if col_od else None,
                "Sched_Weight": df[col_weight] if col_weight else None,
                "Process": df[col_process] if col_process else None,
                "Next": df[col_next] if col_next else None,
                "Pass": df[col_pass] if col_pass else None,
                "Designated_Temp": df[col_heat],
            })
            return out, sheet

    raise ValueError(
        "Couldn't find a sheet in the Ann file with both 'Targeted Th.' and "
        "'HEAT - S.T' columns. Check the file structure."
    )


def build_ready_table(l2_df, ann_df):
    merged = l2_df.merge(ann_df, on="_coil_key", how="inner")
    if merged.empty:
        return merged

    def is_match(row):
        th = row.get("TH_mm")
        if pd.isna(th):
            return True
        return abs(row["Exit_mm"] - th) <= TOLERANCE_MM

    merged["_matched"] = merged.apply(is_match, axis=1)
    return merged[merged["_matched"]].copy()


def style_workbook(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ready for F.Ann"

    headers = [
        "Coil No.", "Customer", "Alloy", "Temper", "Last Pass",
        "Entry Th. (mm)", "Reached Th. (mm)", "Final Targeted Th. (mm)",
        "Width (mm)", "T.W (mm)", "OD (mm)", "Sched. Weight (kg)",
        "Measured Weight (kg)", "Current Process", "Next Process",
        "Rolling End Time", "Avg Speed (m/min)", "Designated Ann. Temp & Time",
    ]

    ws.append(["FINAL ANNEALING - COILS READY (Reached Target Thickness at CRM)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws.append([f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"])
    ws.append([])

    hdr_row = 4
    ws.append(headers)
    title_fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = header_align
        cell.border = border

    r = hdr_row + 1
    for _, row in df.iterrows():
        weight_kg = row["_weight_t"] * 1000 if pd.notna(row.get("_weight_t")) else None
        end_time = row.get("_end_time")
        end_time_str = end_time.strftime("%d-%b-%Y %H:%M") if isinstance(end_time, (datetime,)) else (
            str(end_time) if pd.notna(end_time) else ""
        )
        values = [
            row["_coil_no_display"], row.get("Customer"), row.get("Alloy"), row.get("Temper"),
            row.get("Pass") or row.get("_pass_no"), round(row["Entry_mm"], 2), round(row["Exit_mm"], 2),
            row.get("Targeted_Th_mm"), row.get("Width"), row.get("TW"), row.get("OD"),
            row.get("Sched_Weight"), weight_kg, row.get("Process"), row.get("Next"),
            end_time_str, row.get("_speed"), row.get("Designated_Temp"),
        ]
        ws.append(values)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center
        r += 1

    for row in range(hdr_row + 1, r):
        for col_letter in ["F", "G", "H"]:
            ws[f"{col_letter}{row}"].number_format = "0.00"
        for col_letter in ["L", "M"]:
            ws[f"{col_letter}{row}"].number_format = "#,##0"

    widths = [14, 30, 8, 10, 10, 12, 13, 16, 11, 10, 11, 14, 15, 13, 12, 16, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(headers))}{hdr_row}"
    ws.freeze_panes = f"A{hdr_row + 1}"

    for r_ in ws.iter_rows():
        for c_ in r_:
            c_.font = Font(name="Arial", size=c_.font.size or 10, bold=c_.font.bold, color=c_.font.color)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.title("🔥 Final Annealing — Coils Ready Tool")
st.caption("Upload the L2 machine report + Ann list file. The tool cross-references them and returns a ready styled table for coils that finished CRM and reached their target thickness.")

col1, col2 = st.columns(2)
with col1:
    l2_file = st.file_uploader("L2 Machine Report", type=["xls", "xlsx"])
with col2:
    ann_file = st.file_uploader("Ann List file (with Targeted Th. + HEAT - S.T)", type=["xls", "xlsx"])

tolerance = st.slider("Matching tolerance (mm)", 0.0, 0.10, TOLERANCE_MM, 0.01)
TOLERANCE_MM = tolerance

if l2_file and ann_file:
    if st.button("Process", type="primary"):
        try:
            with st.spinner("Reading L2 report..."):
                l2_df = load_l2_report(l2_file)
            with st.spinner("Reading Ann list..."):
                ann_df, sheet_used = load_ann_target_sheet(ann_file)
            st.success(f"Ann sheet used: **{sheet_used}**")

            ready_df = build_ready_table(l2_df, ann_df)

            if ready_df.empty:
                st.warning("No coils matched — none of the L2 coils reached their target thickness for this stage yet.")
            else:
                st.success(f"{len(ready_df)} coil(s) ready for Final Annealing")
                display_cols = {
                    "_coil_no_display": "Coil No.", "Customer": "Customer", "Alloy": "Alloy",
                    "Temper": "Temper", "Exit_mm": "Reached Th. (mm)",
                    "Targeted_Th_mm": "Final Targeted Th. (mm)", "Designated_Temp": "Designated Temp",
                }
                st.dataframe(ready_df[list(display_cols.keys())].rename(columns=display_cols), use_container_width=True)

                excel_buf = style_workbook(ready_df)
                st.download_button(
                    "⬇️ Download Excel",
                    data=excel_buf,
                    file_name=f"Final_Annealing_Ready_Coils_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Error: {e}")
else:
    st.info("Upload both files to start.")
